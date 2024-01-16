# coding=utf-8
'''
@Time     : 2024/01/15 18:11:38
@Author   : XHao
@Email    : 2510383889@qq.com
'''
# here put the import lib

# 1、利用import导入第三方库openpyxl
from openpyxl import load_workbook
from pathlib import Path


def process(read_location, write_location):
    # 3、加载表格
    # excelRead=load_workbook(file_location)
    # excelWrite=load_workbook(out_location)
    excelRead = load_workbook(read_location)
    excelWrite = load_workbook(write_location)

    # 4、读取sheet页
    tableRead = excelRead['每日统计']
    tableWrite = excelWrite["Sheet1"]

    # 5、获取工作表的行数和列数
    maxRowRead = tableRead.max_row
    maxColumnRead = tableRead.max_column
    maxRowWrite = tableWrite.max_row
    maxColumnWrite = tableWrite.max_column

    # 统计人员的统计时长
    totalHours = float(0.0)
    # 计算天数
    days_num = 0
    firstStuName = tableRead.cell(row=5, column=1).value
    for day_ in range((int)((maxRowRead))):
        stuNameRead = tableRead.cell(row=day_ + 5, column=1).value
        if stuNameRead != firstStuName:
            days_num = day_
            break
    print(f"{days_num}days, {(maxRowRead - 4) / days_num}位同学")
    for rowDing in range((int)((maxRowRead - 4) / days_num)):
        # 获取人名在result表中的所在的行，后续根据行写入数据
        stuNameRead = tableRead.cell(row=rowDing * days_num + 5, column=1).value
        rowWrite = 0
        for stuRow in range(maxRowWrite):
            stuNameWrite = tableWrite.cell(row=stuRow + 1, column=2).value
            if stuNameWrite == stuNameRead:
                rowWrite = stuRow + 1
        if rowWrite == 0:
            continue

        # 单人工作时间统计
        personTotalHours = float(0.0)
        # 每人的周一到周日的时间统计
        for day in range(days_num):
            # stuName = tableRead.cell(row=rowDing * 7 + 5 + day, column=1).value
            # stuGroup = tableRead.cell(row=rowDing * 7 + 5 + day, column=2).value
            totalTime = tableRead.cell(row=rowDing * days_num + 5 + day, column=25).value
            print(totalTime)

            if totalTime != None:
                hourWork = format((int)(totalTime) / 60 + 0.005, '.1f')
            else:
                hourWork = 0

            # 用单元格写入数据
            tableWrite.cell(row=rowWrite, column=6 + day).value = hourWork
            excelWrite.save(write_location)
            personTotalHours += float(hourWork)
        tableWrite.cell(row=rowWrite, column=days_num + 6).value = personTotalHours
        excelWrite.save(write_location)
        totalHours += personTotalHours

    # 完成信息插入之后，进行必要的数据统计
    # 统计完成之后做人均小时数统计，
    tableWrite.cell(row=1, column=days_num + 6).value = totalHours
    excelWrite.save(write_location)
    print(totalHours)


if __name__ == '__main__':
    # 2、打开文件
    root = Path(r"C:\Users\25103\OneDrive\文档\安大研究生\cvpr组\考勤统计\月")
    read_location = root / 'CVPR2023级研究生23.12.01-12.31考勤统计.xlsx'
    write_location = root / 'CVPR2023级研究生23.12.01-12.31考勤统计_结果.xlsx'

    process(read_location, write_location)
