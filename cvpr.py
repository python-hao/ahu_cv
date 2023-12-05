# coding=utf-8
'''
@Time     : 2023/12/04 17:04:24
@Author   : XHao
@Email    : 2510383889@qq.com
'''
# here put the import lib

# 1、利用import导入第三方库openpyxl

from openpyxl import load_workbook
from pathlib import Path

# 2、打开文件
root = Path(r"C:\Users\xhao\OneDrive\文档\安大研究生\cvpr组\考勤统计").resolve()
read_location = root / 'CVPR2023级研究生23.11.27-12.03考勤统计_origin.xlsx'
write_location = root / 'CVPR2023级研究生23.11.27-12.03考勤统计.xlsx'

# 3、加载表格
excelRead = load_workbook(read_location)
excelWrite = load_workbook(write_location)

# 4、读取sheet页
tableRead = excelRead['每日统计']
tableWrite = excelWrite['Sheet1']

# 5、获取工作表的行数和列数
maxRowRead = tableRead.max_row
maxColumnRead = tableRead.max_column
maxRowWrite = tableWrite.max_row
maxColumnWrite = tableWrite.max_column

# 统计人员的统计时长
totalHours = float(0.0)
for rowDing in range((int)((maxRowRead-4)/7)):
    # 获取人名在result表中的所在的行，后续根据行写入数据
    stuNameRead = tableRead.cell(row=rowDing*7+5, column=1).value
    rowWrite = 0
    for stuRow in range(maxRowWrite):
        stuNameWrite = tableWrite.cell(row=stuRow+1, column=2).value
        if stuNameWrite == stuNameRead:
            rowWrite = stuRow+1
    if rowWrite == 0:
        continue

    # 单人工作时间统计
    personTotalHours = float(0.0)
    # 每人的而周一到周日的时间统计
    for rowWeek in range(7):
        stuName = tableRead.cell(row=rowDing*7+5+rowWeek, column=1).value
        stuGroup = tableRead.cell(row=rowDing*7+5+rowWeek, column=2).value
        totalTime = tableRead.cell(row=rowDing*7+5+rowWeek, column=25).value

        if totalTime != None:
            hourWork = format((int)(totalTime)/60+0.005, '.1f')
        else:
            hourWork = 0

        # 用单元格写入数据
        tableWrite.cell(row=rowWrite, column=6+rowWeek).value = hourWork
        excelWrite.save(write_location)
        personTotalHours += float(hourWork)
    tableWrite.cell(row=rowWrite, column=13).value = personTotalHours
    excelWrite.save(write_location)
    totalHours += personTotalHours

# 完成信息插入之后，进行必要的数据统计
# 统计完成之后做人均小时数统计，
tableWrite.cell(row=1, column=33).value = totalHours
excelWrite.save(write_location)
print(totalHours)
